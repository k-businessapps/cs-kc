from __future__ import annotations

import base64
import hmac
import json
import math
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from zoneinfo import ZoneInfo

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

APP_TITLE = "KrispCall Deals Recovery Analyzer"
KTM = ZoneInfo("Asia/Kathmandu")
EMAIL_REGEX = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.IGNORECASE)

KC_PURPLE = "#B04EF0"
KC_DEEP = "#7C3AED"
KC_PINK = "#E060F0"
KC_SOFT = "#F8F2FF"
KC_TEXT = "#1F1430"
KC_GREEN = "#14B86A"
KC_RED = "#E24B5B"

HEADER_FILL = PatternFill("solid", fgColor="EEE2FF")
SUBHEADER_FILL = PatternFill("solid", fgColor="F8F2FF")
METRIC_FILL = PatternFill("solid", fgColor="F6F0FF")
BORDER = Border(bottom=Side(style="thin", color="D8C2F7"))


@dataclass
class PipelineResult:
    pipeline_name: str
    enriched_df: pd.DataFrame
    summary_df: pd.DataFrame
    total_deals: int
    won_deals: int
    deal_value_sum: float
    refund_value_sum: float
    net_value_sum: float
    deduped_from: int
    deduped_to: int


# -----------------------------
# Secrets and auth
# -----------------------------
def get_secret(path: List[str], default=None):
    cur = st.secrets
    for key in path:
        if key not in cur:
            return default
        cur = cur[key]
    return cur


def check_required_secrets() -> List[str]:
    missing = []
    required = [
        (["auth", "username"], "auth.username"),
        (["auth", "password"], "auth.password"),
        (["mixpanel", "project_id"], "mixpanel.project_id"),
        (["mixpanel", "base_url"], "mixpanel.base_url"),
        (["mixpanel", "auth_header"], "mixpanel.auth_header"),
    ]
    for path, label in required:
        if get_secret(path) in (None, ""):
            missing.append(label)
    return missing


def inject_css() -> None:
    st.markdown(
        f"""
        <style>
          .stApp {{ background: linear-gradient(180deg, #ffffff 0%, #fcf9ff 100%); color: {KC_TEXT}; }}
          .block-container {{ padding-top: 1rem; padding-bottom: 1.25rem; max-width: 1280px; }}
          .kc-hero {{
            border-radius: 22px;
            padding: 20px 22px;
            background: linear-gradient(90deg, {KC_DEEP} 0%, {KC_PURPLE} 45%, {KC_PINK} 100%);
            color: white;
            box-shadow: 0 14px 34px rgba(103, 35, 196, 0.18);
          }}
          .kc-hero h1 {{ margin: 0; font-size: 2rem; line-height: 1.1; }}
          .kc-hero p {{ margin: 8px 0 0 0; font-size: 0.96rem; opacity: 0.94; }}
          .kc-panel {{
            background: white;
            border-radius: 18px;
            border: 1px solid rgba(176, 78, 240, 0.14);
            padding: 14px 16px;
            box-shadow: 0 10px 28px rgba(31, 20, 48, 0.05);
          }}
          .kc-chip {{
            display: inline-block;
            padding: 6px 10px;
            border-radius: 999px;
            background: {KC_SOFT};
            color: {KC_DEEP};
            border: 1px solid rgba(176, 78, 240, 0.18);
            font-size: 0.86rem;
            font-weight: 600;
            margin-right: 8px;
            margin-bottom: 8px;
          }}
          .kc-note {{ font-size: 0.9rem; color: rgba(31, 20, 48, 0.78); }}
          div.stButton > button, div.stDownloadButton > button {{
            border-radius: 14px !important;
            border: 0 !important;
            background: linear-gradient(90deg, {KC_DEEP} 0%, {KC_PURPLE} 55%, {KC_PINK} 100%) !important;
            color: white !important;
            font-weight: 700 !important;
            padding: 0.62rem 1rem !important;
            box-shadow: 0 10px 24px rgba(103, 35, 196, 0.18) !important;
          }}
          section[data-testid="stFileUploaderDropzone"] {{
            border-radius: 16px;
            border: 2px dashed rgba(176, 78, 240, 0.35);
            background: {KC_SOFT};
          }}
          div[data-testid="metric-container"] {{
            background: white;
            border: 1px solid rgba(176, 78, 240, 0.14);
            border-radius: 18px;
            padding: 10px 14px;
            box-shadow: 0 10px 28px rgba(31, 20, 48, 0.04);
          }}
          div[data-testid="stDataFrame"] {{
            border-radius: 16px;
            overflow: hidden;
            border: 1px solid rgba(176, 78, 240, 0.12);
          }}
          .stTabs [data-baseweb="tab-list"] {{ gap: 8px; }}
          .stTabs [data-baseweb="tab"] {{
            border-radius: 12px;
            padding: 8px 12px;
            background: white;
            border: 1px solid rgba(176, 78, 240, 0.12);
          }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def logo_html(width_px: int = 245) -> str:
    logo_path = Path(__file__).parent / "assets" / "KrispCallLogo.png"
    if not logo_path.exists():
        return ""
    b64 = base64.b64encode(logo_path.read_bytes()).decode("utf-8")
    return f'<img src="data:image/png;base64,{b64}" style="width:{width_px}px;height:auto;"/>'


def require_login() -> None:
    st.session_state.setdefault("authenticated", False)
    if st.session_state["authenticated"]:
        return

    inject_css()
    missing = check_required_secrets()
    if missing:
        st.error("Missing Streamlit secrets: " + ", ".join(missing))
        st.stop()

    left, right = st.columns([1, 1.6], vertical_alignment="center")
    with left:
        st.markdown(logo_html(280), unsafe_allow_html=True)
    with right:
        st.markdown(
            '<div class="kc-hero"><h1>KrispCall Secure Access</h1><p>Login required before viewing the deal recovery dashboard.</p></div>',
            unsafe_allow_html=True,
        )

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submit = st.form_submit_button("Login")

    if submit:
        expected_user = str(get_secret(["auth", "username"]))
        expected_pass = str(get_secret(["auth", "password"]))
        ok = hmac.compare_digest(username, expected_user) and hmac.compare_digest(password, expected_pass)
        if ok:
            st.session_state["authenticated"] = True
            st.rerun()
        st.error("Invalid username or password.")

    st.stop()


# -----------------------------
# Utility helpers
# -----------------------------
def pick_col(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    lower = {str(c).strip().lower(): c for c in df.columns}
    for candidate in candidates:
        found = lower.get(candidate.strip().lower())
        if found is not None:
            return found
    return None


def extract_first_email(value) -> Optional[str]:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    matches = EMAIL_REGEX.findall(str(value))
    if not matches:
        return None
    return matches[0].strip().lower()


def normalize_email(email: Optional[str]) -> Optional[str]:
    if email is None:
        return None
    found = extract_first_email(email)
    return found if found else None


def parse_deal_created(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce")
    if getattr(parsed.dt, "tz", None) is not None:
        return parsed.dt.tz_convert(KTM).dt.tz_localize(None)
    return parsed


def epoch_series_to_nepal_naive(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        parsed = pd.to_datetime(series, errors="coerce", utc=True)
    else:
        if float(numeric.dropna().median()) > 1e11:
            numeric = numeric / 1000.0
        parsed = pd.to_datetime(numeric, unit="s", errors="coerce", utc=True)
    return parsed.dt.tz_convert(KTM).dt.tz_localize(None)


def money_or_zero(value) -> float:
    num = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return 0.0 if pd.isna(num) else float(num)


def safe_copy(df: pd.DataFrame) -> pd.DataFrame:
    return df.copy() if df is not None else pd.DataFrame()


def canonical_pipeline(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().lower()
    if not text:
        return None
    if "cancel" in text:
        return "Cancelled"
    if "expire" in text:
        return "Expired"
    return None


def format_money(value: float) -> str:
    return f"{value:,.2f}"


def read_uploaded_table(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    if name.endswith(".xlsx") or name.endswith(".xls"):
        return pd.read_excel(uploaded_file)
    raise ValueError("Unsupported file type. Please upload CSV or Excel.")


# -----------------------------
# Mixpanel export
# -----------------------------
def mixpanel_headers() -> Dict[str, str]:
    auth_header = get_secret(["mixpanel", "auth_header"])
    if not auth_header:
        raise RuntimeError("mixpanel.auth_header is missing in Streamlit secrets.")
    return {
        "accept": "text/plain",
        "authorization": str(auth_header).strip(),
    }


@st.cache_data(show_spinner=False, ttl=900)
def fetch_mixpanel_export(
    project_id: str,
    base_url: str,
    event_name: str,
    from_date: date,
    to_date: date,
) -> pd.DataFrame:
    url = f"{str(base_url).rstrip('/')}/api/2.0/export"
    params = {
        "project_id": str(project_id),
        "from_date": from_date.isoformat(),
        "to_date": to_date.isoformat(),
        "event": json.dumps([event_name]),
    }
    response = requests.get(url, params=params, headers=mixpanel_headers(), timeout=240)
    if response.status_code != 200:
        body = (response.text or "")[:600]
        raise RuntimeError(f"Mixpanel export failed for {event_name}. Status {response.status_code}. Response: {body}")

    rows: List[dict] = []
    for line in response.text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError:
            continue

    if not rows:
        return pd.DataFrame()

    raw = pd.DataFrame(rows)
    if "properties" in raw.columns:
        props = pd.json_normalize(raw["properties"])
        raw = pd.concat([raw.drop(columns=["properties"]), props], axis=1)

    if "time" in raw.columns:
        raw["event_time_npt"] = epoch_series_to_nepal_naive(raw["time"])

    return raw


def dedupe_mixpanel_export(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    if df.empty:
        return df.copy(), 0

    work = df.copy()
    required = [c for c in ["event", "distinct_id", "time", "$insert_id"] if c in work.columns]
    if len(required) < 3:
        return work, 0

    sort_cols = [c for c in ["event_time_npt", "mp_processing_time_ms"] if c in work.columns]
    if sort_cols:
        work = work.sort_values(sort_cols, kind="mergesort")

    before = len(work)
    subset = [c for c in ["event", "distinct_id", "time", "$insert_id"] if c in work.columns]
    work = work.drop_duplicates(subset=subset, keep="last")
    return work, before - len(work)


def prep_payment_df(df: pd.DataFrame) -> pd.DataFrame:
    work = safe_copy(df)
    if work.empty:
        return work
    email_col = pick_col(work, ["$email", "email", "Email", "user.email", "User Email"])
    amount_col = pick_col(work, ["Amount", "amount", "amount_value"])
    desc_col = pick_col(work, ["Amount Description", "description", "Description", "$description"])

    work["email"] = work[email_col].apply(extract_first_email) if email_col else None
    work["Amount"] = pd.to_numeric(work[amount_col], errors="coerce") if amount_col else pd.NA
    work["Amount Description"] = work[desc_col].astype(str) if desc_col else ""
    work = work[work["email"].notna()].copy()
    work = work.sort_values(["email", "event_time_npt"], kind="mergesort")
    return work


def prep_refund_df(df: pd.DataFrame) -> pd.DataFrame:
    work = safe_copy(df)
    if work.empty:
        return work
    email_col = pick_col(work, ["User Email", "$email", "email", "Email", "user.email"])
    amount_col = pick_col(work, ["Refund Amount", "refund_amount", "Amount", "amount"])
    desc_col = pick_col(work, ["Refunded Transaction Description", "Refunded Transaction description", "description"])

    work["email"] = work[email_col].apply(extract_first_email) if email_col else None
    work["Refund Amount"] = pd.to_numeric(work[amount_col], errors="coerce") if amount_col else pd.NA
    work["Refunded Transaction Description"] = work[desc_col].astype(str) if desc_col else ""
    work = work[work["email"].notna()].copy()
    work = work.sort_values(["email", "event_time_npt"], kind="mergesort")
    return work


# -----------------------------
# Deal processing
# -----------------------------
def standardize_deals(deals_df: pd.DataFrame) -> Tuple[pd.DataFrame, Dict[str, str]]:
    work = deals_df.copy()

    created_col = pick_col(work, ["Deal - Deal created on", "Deal created on", "Created on", "Deal Created On"])
    email_col = pick_col(work, ["Person - Email", "Person Email", "Email"])
    title_col = pick_col(work, ["Deal - Title", "Deal Title", "Title"])
    owner_col = pick_col(work, ["Deal - Owner", "Deal Owner", "Owner"])
    pipeline_col = pick_col(work, ["Deal - Pipeline", "Deal Pipeline", "Pipeline"])
    status_col = pick_col(work, ["Deal - Status", "Deal Status", "Status"])

    missing = {
        "Deal - Deal created on": created_col,
        "Deal - Owner": owner_col,
        "Deal - Pipeline": pipeline_col,
    }
    missing_required = [label for label, col in missing.items() if col is None]
    if missing_required:
        raise ValueError("Missing required columns: " + ", ".join(missing_required))

    work["Deal_Created_NPT"] = parse_deal_created(work[created_col])
    work["Person_Email_Extracted"] = work[email_col].apply(extract_first_email) if email_col else None
    work["Deal_Title_Email_Extracted"] = work[title_col].apply(extract_first_email) if title_col else None
    work["Unified_Email"] = work["Person_Email_Extracted"].combine_first(work["Deal_Title_Email_Extracted"])
    work["Pipeline_Group"] = work[pipeline_col].apply(canonical_pipeline)
    work["Deal_Owner"] = work[owner_col].astype(str).str.strip()
    work["Deal_Status_Normalized"] = work[status_col].astype(str).str.strip() if status_col else ""

    mapping = {
        "created_col": created_col,
        "email_col": email_col or "",
        "title_col": title_col or "",
        "owner_col": owner_col,
        "pipeline_col": pipeline_col,
        "status_col": status_col or "",
    }
    return work, mapping


def dedupe_pipeline(df: pd.DataFrame) -> Tuple[pd.DataFrame, int, int]:
    if df.empty:
        return df.copy(), 0, 0
    before = len(df)
    deduped = (
        df[df["Unified_Email"].notna()].copy()
        .sort_values(["Unified_Email", "Deal_Created_NPT"], kind="mergesort")
        .drop_duplicates(subset=["Unified_Email"], keep="first")
        .copy()
    )
    return deduped, before, len(deduped)


def build_event_maps(payments_df: pd.DataFrame, refunds_df: pd.DataFrame) -> Tuple[Dict[str, pd.DataFrame], Dict[str, pd.DataFrame]]:
    pay_map = {email: grp.sort_values("event_time_npt", kind="mergesort").reset_index(drop=True) for email, grp in payments_df.groupby("email", sort=False)} if not payments_df.empty else {}
    refund_map = {email: grp.sort_values("event_time_npt", kind="mergesort").reset_index(drop=True) for email, grp in refunds_df.groupby("email", sort=False)} if not refunds_df.empty else {}
    return pay_map, refund_map


def enrich_pipeline(df: pd.DataFrame, pay_map: Dict[str, pd.DataFrame], refund_map: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    records = []
    for _, row in df.iterrows():
        email = row.get("Unified_Email")
        created = row.get("Deal_Created_NPT")

        deal_value = math.nan
        first_payment_time = pd.NaT
        first_payment_desc = None
        won = False
        refund_value = 0.0

        if pd.notna(created) and email in pay_map:
            p_df = pay_map[email]
            p_match = p_df[p_df["event_time_npt"] > created]
            if not p_match.empty:
                first = p_match.iloc[0]
                deal_value = money_or_zero(first.get("Amount"))
                first_payment_time = first.get("event_time_npt")
                first_payment_desc = first.get("Amount Description")
                won = True

        if won and pd.notna(created) and email in refund_map:
            r_df = refund_map[email]
            window_end = created + timedelta(days=15)
            r_match = r_df[(r_df["event_time_npt"] > created) & (r_df["event_time_npt"] <= window_end)]
            refund_value = float(pd.to_numeric(r_match.get("Refund Amount", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())

        out = row.to_dict()
        out["Deal_Value"] = None if pd.isna(deal_value) else float(deal_value)
        out["Refund_Value"] = float(refund_value)
        out["Won"] = bool(won)
        out["Net_Value"] = (0.0 if pd.isna(deal_value) else float(deal_value)) - float(refund_value)
        out["First_Payment_Time_NPT"] = first_payment_time
        out["First_Payment_Description"] = first_payment_desc
        records.append(out)

    enriched = pd.DataFrame(records)
    if not enriched.empty:
        enriched["Won"] = enriched["Won"].fillna(False)
        enriched["Refund_Value"] = pd.to_numeric(enriched["Refund_Value"], errors="coerce").fillna(0.0)
        enriched["Net_Value"] = pd.to_numeric(enriched["Net_Value"], errors="coerce").fillna(0.0)
        enriched["Deal_Value"] = pd.to_numeric(enriched["Deal_Value"], errors="coerce")
    return enriched


def build_summary(enriched_df: pd.DataFrame) -> pd.DataFrame:
    if enriched_df.empty:
        return pd.DataFrame(columns=["Deal_Owner", "Won_Count", "Deal_Value", "Refund_Value", "Net_Value"])

    summary = (
        enriched_df.groupby("Deal_Owner", dropna=False)
        .agg(
            Won_Count=("Won", lambda s: int(pd.Series(s).fillna(False).astype(bool).sum())),
            Deal_Value=("Deal_Value", lambda s: float(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
            Refund_Value=("Refund_Value", lambda s: float(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
            Net_Value=("Net_Value", lambda s: float(pd.to_numeric(s, errors="coerce").fillna(0).sum())),
        )
        .reset_index()
        .sort_values(["Net_Value", "Won_Count"], ascending=[False, False], kind="mergesort")
    )
    return summary


def pipeline_result(name: str, source_df: pd.DataFrame, pay_map: Dict[str, pd.DataFrame], refund_map: Dict[str, pd.DataFrame]) -> PipelineResult:
    deduped, before, after = dedupe_pipeline(source_df)
    enriched = enrich_pipeline(deduped, pay_map, refund_map)
    summary = build_summary(enriched)
    won_count = int(enriched["Won"].fillna(False).astype(bool).sum()) if not enriched.empty else 0
    deal_sum = float(pd.to_numeric(enriched.get("Deal_Value", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not enriched.empty else 0.0
    refund_sum = float(pd.to_numeric(enriched.get("Refund_Value", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not enriched.empty else 0.0
    net_sum = float(pd.to_numeric(enriched.get("Net_Value", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()) if not enriched.empty else 0.0
    return PipelineResult(
        pipeline_name=name,
        enriched_df=enriched,
        summary_df=summary,
        total_deals=len(enriched),
        won_deals=won_count,
        deal_value_sum=deal_sum,
        refund_value_sum=refund_sum,
        net_value_sum=net_sum,
        deduped_from=before,
        deduped_to=after,
    )


def run_analysis(
    deals_df: pd.DataFrame,
    payments_df: pd.DataFrame,
    refunds_df: pd.DataFrame,
) -> Tuple[PipelineResult, PipelineResult, List[str]]:
    logs: List[str] = []

    deals_std, mapping = standardize_deals(deals_df)
    logs.append(f"Deals rows loaded: {len(deals_std):,}")
    unified_missing = int(deals_std["Unified_Email"].isna().sum())
    if unified_missing:
        logs.append(f"Deals with no usable email from Person - Email or Deal Title: {unified_missing:,}")

    payments_prepped = prep_payment_df(payments_df)
    refunds_prepped = prep_refund_df(refunds_df)
    logs.append(f"Payment events usable after cleanup: {len(payments_prepped):,}")
    logs.append(f"Refund events usable after cleanup: {len(refunds_prepped):,}")

    pay_map, refund_map = build_event_maps(payments_prepped, refunds_prepped)

    cancelled_source = deals_std[deals_std["Pipeline_Group"] == "Cancelled"].copy()
    expired_source = deals_std[deals_std["Pipeline_Group"] == "Expired"].copy()
    logs.append(f"Cancelled pipeline rows before dedupe: {len(cancelled_source):,}")
    logs.append(f"Expired pipeline rows before dedupe: {len(expired_source):,}")

    cancelled = pipeline_result("Cancelled", cancelled_source, pay_map, refund_map)
    expired = pipeline_result("Expired", expired_source, pay_map, refund_map)

    logs.append(f"Cancelled dedupe kept {cancelled.deduped_to:,} of {cancelled.deduped_from:,} rows")
    logs.append(f"Expired dedupe kept {expired.deduped_to:,} of {expired.deduped_from:,} rows")
    logs.append(
        "Refunds are counted only for won deals, using refunds that happened within 15 days after the deal creation timestamp in Nepal time."
    )
    logs.append(
        f"Detected columns. Created: {mapping['created_col']}, Owner: {mapping['owner_col']}, Pipeline: {mapping['pipeline_col']}"
    )
    if mapping.get("email_col"):
        logs.append(f"Primary email column used: {mapping['email_col']}")
    if mapping.get("title_col"):
        logs.append(f"Fallback title column used for email extraction: {mapping['title_col']}")

    return cancelled, expired, logs


# -----------------------------
# Excel export
# -----------------------------
def display_columns(enriched_df: pd.DataFrame) -> List[str]:
    preferred = [
        "Unified_Email",
        "Deal_Created_NPT",
        "Deal_Owner",
        "Pipeline_Group",
        "Deal_Status_Normalized",
        "Deal_Value",
        "Refund_Value",
        "Won",
        "Net_Value",
        "First_Payment_Time_NPT",
        "First_Payment_Description",
    ]
    keep = [c for c in preferred if c in enriched_df.columns]
    other_core = [
        c for c in [
            "Person_Email_Extracted",
            "Deal_Title_Email_Extracted",
        ] if c in enriched_df.columns and c not in keep
    ]
    passthrough = [c for c in enriched_df.columns if c not in keep + other_core]
    return keep + other_core + passthrough


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="4C1D95")
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", horizontal="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (datetime, pd.Timestamp)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells[:300]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(12, max_len + 2), 38)


def append_df(ws, df: pd.DataFrame) -> None:
    if df.empty:
        ws.append(["No data"])
        return
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        values = []
        for value in row:
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime()
            elif pd.isna(value):
                value = None
            values.append(value)
        ws.append(values)
    format_sheet(ws)


def summary_with_total(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    total = {
        "Deal_Owner": "TOTAL",
        "Won_Count": int(pd.to_numeric(df["Won_Count"], errors="coerce").fillna(0).sum()),
        "Deal_Value": float(pd.to_numeric(df["Deal_Value"], errors="coerce").fillna(0).sum()),
        "Refund_Value": float(pd.to_numeric(df["Refund_Value"], errors="coerce").fillna(0).sum()),
        "Net_Value": float(pd.to_numeric(df["Net_Value"], errors="coerce").fillna(0).sum()),
    }
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


def build_workbook(cancelled: PipelineResult, expired: PipelineResult, logs: List[str]) -> bytes:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    cancel_enriched_ws = wb.create_sheet("Cancelled_Enriched")
    append_df(cancel_enriched_ws, cancelled.enriched_df[display_columns(cancelled.enriched_df)] if not cancelled.enriched_df.empty else cancelled.enriched_df)

    expired_enriched_ws = wb.create_sheet("Expired_Enriched")
    append_df(expired_enriched_ws, expired.enriched_df[display_columns(expired.enriched_df)] if not expired.enriched_df.empty else expired.enriched_df)

    cancel_summary_ws = wb.create_sheet("Cancelled_Summary")
    append_df(cancel_summary_ws, summary_with_total(cancelled.summary_df))

    expired_summary_ws = wb.create_sheet("Expired_Summary")
    append_df(expired_summary_ws, summary_with_total(expired.summary_df))

    for ws in [cancel_summary_ws, expired_summary_ws]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if str(row[0].value).strip().upper() == "TOTAL":
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = SUBHEADER_FILL

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# -----------------------------
# UI
# -----------------------------
def render_metric_row(result: PipelineResult) -> None:
    c1, c2, c3, c4 = st.columns(4)
    c1.metric(f"{result.pipeline_name} deals", f"{result.total_deals:,}")
    c2.metric(f"{result.pipeline_name} won", f"{result.won_deals:,}")
    c3.metric(f"{result.pipeline_name} deal value", format_money(result.deal_value_sum))
    c4.metric(f"{result.pipeline_name} net value", format_money(result.net_value_sum))


def render_pipeline_tab(result: PipelineResult) -> None:
    st.markdown(
        f'<div class="kc-panel"><div class="kc-chip">Deduped {result.deduped_from:,} → {result.deduped_to:,}</div>'
        f'<div class="kc-chip">Won deals: {result.won_deals:,}</div>'
        f'<div class="kc-chip">Refund total: {format_money(result.refund_value_sum)}</div></div>',
        unsafe_allow_html=True,
    )
    st.markdown("#### Enriched rows")
    st.dataframe(result.enriched_df[display_columns(result.enriched_df)] if not result.enriched_df.empty else result.enriched_df, use_container_width=True, hide_index=True)
    st.markdown("#### Summary by Deal Owner")
    st.dataframe(summary_with_total(result.summary_df), use_container_width=True, hide_index=True)


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="📊", layout="wide")
    require_login()
    inject_css()

    st.markdown('<div style="height: 4px;"></div>', unsafe_allow_html=True)
    left, right = st.columns([1, 2.6], vertical_alignment="center")
    with left:
        st.markdown(logo_html(260), unsafe_allow_html=True)
    with right:
        st.markdown(
            '<div class="kc-hero"><h1>KrispCall Deals Recovery Analyzer</h1>'
            '<p>Upload the deals file once. The app pulls <strong>New Payment Made</strong> and <strong>Refund Granted</strong> directly from Mixpanel, one after the other, then builds pipeline-level outputs for Cancelled and Expired deals.</p></div>',
            unsafe_allow_html=True,
        )

    with st.sidebar:
        st.markdown("### Mixpanel export range")
        today = datetime.now(KTM).date()
        default_from = today.replace(day=1)
        from_date = st.date_input("From date", value=default_from)
        to_date = st.date_input("To date", value=today)
        if from_date > to_date:
            st.error("From date cannot be later than To date.")
            st.stop()

        st.markdown("---")
        st.markdown("### Logic")
        st.markdown('<div class="kc-note">Email unification uses <strong>Person - Email</strong> first, then extracts email from <strong>Deal Title</strong> if needed.</div>', unsafe_allow_html=True)
        st.markdown('<div class="kc-note">Each pipeline is deduped by unified email. Current rule keeps the earliest deal created timestamp per email inside that pipeline.</div>', unsafe_allow_html=True)
        st.markdown('<div class="kc-note">Deal Value is the first payment after deal creation. Refund Value is the sum of refunds within 15 days after deal creation, only when the deal is won.</div>', unsafe_allow_html=True)

    st.markdown('<div class="kc-panel">', unsafe_allow_html=True)
    deals_file = st.file_uploader("Upload deals file", type=["csv", "xlsx", "xls"], help="Deals file is the primary input. Mixpanel events are fetched automatically.")
    run = st.button("Run analysis", type="primary", disabled=deals_file is None)
    st.markdown('</div>', unsafe_allow_html=True)

    if not run:
        return

    if deals_file is None:
        st.warning("Upload the deals file to continue.")
        return

    progress = st.progress(0, text="Starting analysis...")
    status = st.empty()

    try:
        status.info("Reading deals file...")
        deals_df = read_uploaded_table(deals_file)
        progress.progress(10, text="Deals file loaded")

        project_id = str(get_secret(["mixpanel", "project_id"]))
        base_url = str(get_secret(["mixpanel", "base_url"]))

        status.info("Fetching Mixpanel event 1 of 2: New Payment Made")
        payments_raw = fetch_mixpanel_export(project_id, base_url, "New Payment Made", from_date, to_date)
        payments_deduped, payments_removed = dedupe_mixpanel_export(payments_raw)
        progress.progress(35, text="New Payment Made fetched")

        status.info("Fetching Mixpanel event 2 of 2: Refund Granted")
        refunds_raw = fetch_mixpanel_export(project_id, base_url, "Refund Granted", from_date, to_date)
        refunds_deduped, refunds_removed = dedupe_mixpanel_export(refunds_raw)
        progress.progress(60, text="Refund Granted fetched")

        status.info("Processing deals, deduplication, and enrichment...")
        cancelled, expired, logs = run_analysis(deals_df, payments_deduped, refunds_deduped)
        logs.insert(0, f"Payments dedupe removed: {payments_removed:,}")
        logs.insert(1, f"Refunds dedupe removed: {refunds_removed:,}")
        logs.insert(2, f"Mixpanel date range used: {from_date.isoformat()} to {to_date.isoformat()}")
        progress.progress(85, text="Enrichment complete")

        workbook_bytes = build_workbook(cancelled, expired, logs)
        filename = f"krispcall_deals_recovery_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"
        progress.progress(100, text="Ready")
        status.success("Analysis complete.")

    except Exception as exc:
        progress.empty()
        status.error(f"Run failed: {exc}")
        st.exception(exc)
        return

    overview_tab, cancelled_tab, expired_tab, export_tab, logs_tab = st.tabs(
        ["Overview", "Cancelled", "Expired", "Export", "Logs"]
    )

    with overview_tab:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Payments events", f"{len(payments_deduped):,}")
        c2.metric("Refund events", f"{len(refunds_deduped):,}")
        c3.metric("Cancelled net", format_money(cancelled.net_value_sum))
        c4.metric("Expired net", format_money(expired.net_value_sum))
        render_metric_row(cancelled)
        render_metric_row(expired)
        st.markdown(
            '<div class="kc-panel"><span class="kc-chip">Cancelled summary tab included</span><span class="kc-chip">Expired summary tab included</span><span class="kc-chip">Run log included in export</span></div>',
            unsafe_allow_html=True,
        )

    with cancelled_tab:
        render_pipeline_tab(cancelled)

    with expired_tab:
        render_pipeline_tab(expired)

    with export_tab:
        st.markdown("#### Download output workbook")
        st.download_button(
            label="Download Excel workbook",
            data=workbook_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.markdown("#### Included sheets")
        st.markdown(
            '<div class="kc-panel"><span class="kc-chip">Cancelled_Enriched</span><span class="kc-chip">Expired_Enriched</span><span class="kc-chip">Cancelled_Summary</span><span class="kc-chip">Expired_Summary</span></div>',
            unsafe_allow_html=True,
        )

    with logs_tab:
        st.dataframe(pd.DataFrame({"log": logs}), use_container_width=True, hide_index=True)


if __name__ == "__main__":
    main()
